import datetime
import pathlib
import tempfile

from anymail.message import AnymailMessage
from django.conf import settings
from django.core.exceptions import ValidationError
from django.core.management import BaseCommand, CommandError
from django.core.validators import validate_email
from django.utils import timezone
from django.utils.dateparse import parse_date
from mtp_common.tasks import default_from_address
import openpyxl
from openpyxl.cell import WriteOnlyCell
from openpyxl.utils import get_column_letter

from credit.constants import CREDIT_STATUS
from credit.models import Credit, LOG_ACTIONS as CREDIT_LOG_ACTIONS
from disbursement.constants import DISBURSEMENT_METHOD, DISBURSEMENT_RESOLUTION
from disbursement.models import Disbursement, LOG_ACTIONS as DISBURSEMENT_LOG_ACTIONS
from notification.rules import RULES, CountingRule, MonitoredRule, Triggered
from transaction.utils import format_amount


class Command(BaseCommand):
    def add_arguments(self, parser):
        super().add_arguments(parser)
        parser.add_argument('--since', help='Since date (inclusive)')
        parser.add_argument('--until', help='Until date (exclusive)')
        parser.add_argument('--rules', nargs='*', choices=RULES.keys(), help='Notification rule codes')
        parser.add_argument('emails', nargs='+', help='Email addresses to send reports to')

    def handle(self, **options):
        emails = options['emails']
        for email in emails:
            try:
                validate_email(email)
            except ValidationError:
                raise CommandError('Email is not valid email address', {'email': email})

        codes = options['rules'] or RULES.keys()
        rules = [RULES[code] for code in codes]

        period_start, period_end = report_period(options['since'], options['until'])
        period_end_inclusize = period_end - datetime.timedelta(days=1)
        if period_start == period_end_inclusize:
            period_filename = period_start.date().isoformat()
            period_description = period_start.strftime('%d %b %Y')
        else:
            period_filename = f'{period_start.date().isoformat()}-{period_end_inclusize.date().isoformat()}'
            period_description = f"{period_start.strftime('%d %b %Y')} to {period_end_inclusize.strftime('%d %b %Y')}"

        with tempfile.TemporaryDirectory() as temp_path:
            temp_path = pathlib.Path(temp_path)
            report_path = temp_path / f'{period_filename}.xlsx'
            workbook = openpyxl.Workbook(write_only=True)
            generate_report(workbook, period_start, period_end, rules)
            workbook.save(report_path)
            send_report(period_description, report_path, emails)

        self.stdout.write('Emailed report')


def make_local_datetime(date):
    return timezone.make_aware(datetime.datetime.combine(date, datetime.time.min))


def report_period(period_start, period_end):
    if period_start:
        period_start = parse_date(period_start)
        if not period_start:
            raise CommandError('Cannot parse `--since`')
        period_start = make_local_datetime(period_start)
    if period_end:
        period_end = parse_date(period_end)
        if not period_end:
            raise CommandError('Cannot parse `--until`')
        period_end = make_local_datetime(period_end)

    today = make_local_datetime(timezone.now())
    if not period_start and not period_end:
        # default is "last week"
        period_end = today - datetime.timedelta(days=today.weekday())
        period_start = period_end - datetime.timedelta(days=7)
    elif not period_start:
        # if only end date is specified, take one day
        period_start = period_end - datetime.timedelta(days=1)
    elif not period_end:
        # if only start date is specified, take period up to today
        period_end = today

    if period_start >= period_end:
        raise CommandError('`--until` must be after `--since`')
    elif period_end > period_start + datetime.timedelta(days=7):
        raise CommandError('Maximum date span is 7 days')

    return period_start, period_end


def generate_report(workbook, period_start, period_end, rules):
    candidate_credits = Credit.objects.filter(
        prisoner_profile__isnull=False,
        sender_profile__isnull=False,
    ).filter(
        received_at__gte=period_start,
        received_at__lt=period_end,
    ).order_by('pk')
    candidate_disbursements = Disbursement.objects.filter(
        prisoner_profile__isnull=False,
        recipient_profile__isnull=False,
        resolution=DISBURSEMENT_RESOLUTION.SENT,
    ).filter(
        created__gte=period_start,
        created__lt=period_end,
    ).order_by('pk')
    records = {
        Credit: candidate_credits,
        Disbursement: candidate_disbursements,
    }

    for rule in rules:
        for serialised_model, serialiser_cls in Serialiser.serialisers.items():
            if serialised_model not in rule.applies_to_models:
                continue
            serialiser = serialiser_cls(rule)
            worksheet = workbook.create_sheet(
                title=f'{serialised_model._meta.verbose_name[:4]}-{rule.abbr_description}'
            )
            generate_sheet(worksheet, serialiser, rule, records[serialised_model])


def generate_sheet(worksheet, serialiser, rule, record_set):
    headers = serialiser.get_headers()
    worksheet.append(headers)
    count = 0
    for record in record_set:
        if not rule.applies_to(record):
            continue
        triggered = rule.triggered(record)
        if not triggered:
            continue
        row = serialiser.serialise(worksheet, record, triggered)
        worksheet.append([
            row.get(field, None)
            for field in headers
        ])
        count += 1
    if count:
        worksheet.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{count + 1}'
    else:
        note = WriteOnlyCell(worksheet, 'No notifications')
        note.style = 'Good'
        worksheet.append([serialiser.rule_description, note])


def send_report(period_description, report_path, emails):
    email = AnymailMessage(
        subject=f'Prisoner money notifications for {period_description}',
        body=f"""
OFFICIAL SENSITIVE

Please find attached, the prisoner money notifications report for {period_description}.

There is a separate sheet for each notification rule for credits and disbursements.

The ‘Monitored by’ column that appears in some sheets is the number of users
who are monitoring that prisoner or payment source.

The ‘How many?’ column that appears in some sheets is the number that triggered
the rule in column A. For example, if the ‘How many?’ column says 4 for the rule
‘More than 2 credits from the same debit card or bank account to any prisoner in a week’,
then this means that a specific debit card or bank account sent 4 credits in a week
up to when that credit was sent.

If you have any queries, contact the team at {settings.TEAM_EMAIL}.
        """.strip(),
        from_email=default_from_address(),
        to=emails,
        tags=['notifications-report'],
    )
    email.attach_file(str(report_path), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    email.send()


class Serialiser:
    serialisers = {}
    additional_headers = {
        MonitoredRule: {'header': 'Monitored by', 'triggered_kwarg': 'monitoring_user_count'},
        CountingRule: {'header': 'How many?', 'triggered_kwarg': 'count'},
    }

    def __init_subclass__(cls, serialised_model):
        cls.serialisers[serialised_model] = cls

    def __init__(self, rule):
        self.rule = rule

    @property
    def rule_description(self):
        return self.rule.description.replace('you are monitoring', 'someone monitors')

    def get_headers(self):
        headers = [
            'Notification rule',
            'Internal ID',
        ]
        additional_header = self.additional_headers.get(type(self.rule), None)
        if additional_header:
            headers.insert(1, additional_header['header'])
        return headers

    def serialise(self, worksheet, record, triggered: Triggered):
        linked_cell = WriteOnlyCell(worksheet, self.get_internal_id(record))
        linked_cell.hyperlink = self.get_noms_ops_url(record)
        linked_cell.style = 'Hyperlink'
        row = {
            'Notification rule': self.rule_description,
            'Internal ID': linked_cell,
        }
        additional_header = self.additional_headers.get(type(self.rule), None)
        if additional_header:
            row[additional_header['header']] = triggered.kwargs[additional_header['triggered_kwarg']]
        return row

    def get_internal_id(self, record):
        raise NotImplementedError

    def get_noms_ops_url(self, record):
        raise NotImplementedError


class CreditSerialiser(Serialiser, serialised_model=Credit):
    def get_headers(self):
        return super().get_headers() + [
            'Date started', 'Date received', 'Date credited',
            'Amount',
            'Prisoner number', 'Prisoner name', 'Prison',
            'Sender name', 'Payment method',
            'Bank transfer sort code', 'Bank transfer account', 'Bank transfer roll number',
            'Debit card number', 'Debit card expiry', 'Debit card billing address',
            'Sender email', 'Sender IP address',
            'Status',
            'NOMIS transaction',
            'WorldPay order code',
        ]

    def serialise(self, worksheet, record: Credit, triggered: Triggered):
        row = super().serialise(worksheet, record, triggered)
        status = record.status
        if status:
            status = str(CREDIT_STATUS.for_value(status).display)
        else:
            status = 'Anonymous'
        row.update({
            'Date received': local_datetime_for_xlsx(record.received_at),
            'Date credited': local_datetime_for_xlsx(record.log_set.get_action_date(CREDIT_LOG_ACTIONS.CREDITED)),
            'Amount': format_amount(record.amount),
            'Prisoner number': record.prisoner_number or 'Unknown',
            'Prisoner name': record.prisoner_name or 'Unknown',
            'Prison': record.prison.short_name if record.prison else 'Unknown',
            'Status': status,
            'NOMIS transaction': record.nomis_transaction_id,
            **self.serialise_sender(record)
        })
        return row

    def get_internal_id(self, record):
        return f'Credit {record.id}'

    def get_noms_ops_url(self, record):
        return f'{settings.NOMS_OPS_URL}/security/credits/{record.id}/'

    def serialise_sender(self, record: Credit):
        if hasattr(record, 'transaction'):
            transaction = record.transaction
            return {
                'Payment method': 'Bank transfer',
                'Sender name': transaction.sender_name,
                'Bank transfer sort code': transaction.sender_sort_code,
                'Bank transfer account': transaction.sender_account_number,
                'Bank transfer roll number': transaction.sender_roll_number,
            }

        if hasattr(record, 'payment'):
            payment = record.payment
            return {
                'Date started': local_datetime_for_xlsx(payment.created),
                'Payment method': 'Debit card',
                'Sender name': payment.cardholder_name,
                'Debit card number': (
                    f'{payment.card_number_first_digits or "******"}******{payment.card_number_last_digits}'
                ),
                'Debit card expiry': payment.card_expiry_date,
                'Debit card billing address': str(payment.billing_address),
                'Sender email': payment.email,
                'Sender IP address': payment.ip_address,
                'WorldPay order code': payment.worldpay_id,
            }

        return {
            'Payment method': 'Unknown',
            'Sender name': '(Unknown sender)',
        }


class DisbursementSerialiser(Serialiser, serialised_model=Disbursement):
    def get_headers(self):
        return super().get_headers() + [
            'Date entered', 'Date confirmed', 'Date sent',
            'Amount',
            'Prisoner number', 'Prisoner name', 'Prison',
            'Recipient name', 'Payment method',
            'Bank transfer sort code', 'Bank transfer account', 'Bank transfer roll number',
            'Recipient address', 'Recipient email',
            'Status',
            'NOMIS transaction', 'SOP invoice number',
        ]

    def serialise(self, worksheet, record: Disbursement, triggered: Triggered):
        row = super().serialise(worksheet, record, triggered)
        row.update({
            'Date entered': local_datetime_for_xlsx(record.created),
            'Date confirmed': local_datetime_for_xlsx(
                record.log_set.get_action_date(DISBURSEMENT_LOG_ACTIONS.CONFIRMED)
            ),
            'Date sent': local_datetime_for_xlsx(record.log_set.get_action_date(DISBURSEMENT_LOG_ACTIONS.SENT)),
            'Amount': format_amount(record.amount),
            'Prisoner number': record.prisoner_number,
            'Prisoner name': record.prisoner_name,
            'Prison': record.prison.short_name,
            'Recipient name': record.recipient_name,
            'Payment method': str(DISBURSEMENT_METHOD.for_value(record.method).display),
            'Bank transfer sort code': record.sort_code,
            'Bank transfer account': record.account_number,
            'Bank transfer roll number': record.roll_number,
            'Recipient address': record.recipient_address,
            'Recipient email': record.recipient_email,
            'Status': str(DISBURSEMENT_RESOLUTION.for_value(record.resolution).display),
            'NOMIS transaction': record.nomis_transaction_id,
            'SOP invoice number': record.invoice_number,
        })
        return row

    def get_internal_id(self, record):
        return f'Disbursement {record.id}'

    def get_noms_ops_url(self, record):
        return f'{settings.NOMS_OPS_URL}/security/disbursements/{record.id}/'


def local_datetime_for_xlsx(value):
    if not value:
        return None
    return timezone.make_naive(value)
