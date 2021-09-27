import datetime
import io
from unittest import mock

from django.core.management import CommandError, call_command
from django.test import TestCase
from django.utils import timezone
from model_mommy import mommy
import openpyxl
from openpyxl.utils import coordinate_to_tuple

from core.tests.utils import make_test_users, FLAKY_TEST_WARNING
from credit.constants import CREDIT_RESOLUTION
from credit.models import Credit
from disbursement.models import Disbursement
from disbursement.tests.utils import generate_disbursements
from notification.constants import EMAIL_FREQUENCY
from notification.management.commands.send_notification_emails import (
    EMAILS_STARTED_FLAG,
    get_events, group_events, summarise_group,
)
from notification.models import Event, EmailNotificationPreferences
from notification.rules import RULES
from notification.tests.utils import make_sender, make_prisoner, make_csfreq_credits
from payment.constants import PAYMENT_STATUS
from payment.models import Payment
from payment.tests.utils import generate_payments
from prison.models import PrisonerLocation
from prison.tests.utils import load_random_prisoner_locations
from security.models import PrisonerProfile, SenderProfile, DebitCardSenderDetails


class NotificationBaseTestCase(TestCase):
    fixtures = ['initial_types.json', 'test_prisons.json', 'initial_groups.json']

    def create_profiles_but_unlink_objects(self):
        call_command('update_security_profiles')
        Credit.objects.update(
            is_counted_in_sender_profile_total=False,
            is_counted_in_prisoner_profile_total=False
        )
        Disbursement.objects.update(recipient_profile=None, prisoner_profile=None)
        # NB: profiles will have incorrect counts and totals


@mock.patch('notification.management.commands.send_notification_emails.send_email')
class SendNotificationEmailsTestCase(NotificationBaseTestCase):
    def setUp(self):
        super().setUp()
        test_users = make_test_users()
        self.security_staff = test_users['security_staff']
        load_random_prisoner_locations()
        generate_payments(
            payment_batch=20, days_of_history=2,
            overrides={'status': PAYMENT_STATUS.TAKEN, 'credited': True}
        )
        generate_disbursements(disbursement_batch=20, days_of_history=1)

    def test_does_not_send_email_notifications_for_no_events(self, mock_send_email):
        user = self.security_staff[0]
        user.flags.create(name=EMAILS_STARTED_FLAG)
        EmailNotificationPreferences(user=user, frequency=EMAIL_FREQUENCY.DAILY).save()
        call_command('send_notification_emails')

        self.assertFalse(Event.objects.exists())
        mock_send_email.assert_not_called()
        self.assertIsNone(EmailNotificationPreferences.objects.get(user=user).last_sent_at)

    def test_does_not_send_email_notifications_for_no_monitoring(self, mock_send_email):
        user = self.security_staff[0]
        user.flags.create(name=EMAILS_STARTED_FLAG)
        EmailNotificationPreferences(user=user, frequency=EMAIL_FREQUENCY.DAILY).save()
        call_command('update_security_profiles')
        call_command('send_notification_emails')

        self.assertFalse(Event.objects.exists())
        mock_send_email.assert_not_called()
        self.assertIsNone(EmailNotificationPreferences.objects.get(user=user).last_sent_at)

    def test_sends_first_email_not_monitoring(self, mock_send_email):
        user = self.security_staff[0]
        EmailNotificationPreferences(user=user, frequency=EMAIL_FREQUENCY.DAILY).save()
        call_command('update_security_profiles')
        call_command('send_notification_emails')

        self.assertEqual(len(mock_send_email.call_args_list), 1)
        send_email_kwargs = mock_send_email.call_args_list[0].kwargs
        self.assertEqual(send_email_kwargs['template_name'], 'api-intel-notification-not-monitoring')
        self.assertTrue(user.flags.filter(name=EMAILS_STARTED_FLAG).exists())
        self.assertIsNotNone(EmailNotificationPreferences.objects.get(user=user).last_sent_at)

    def test_sends_first_email_with_events(self, mock_send_email):
        user = self.security_staff[0]
        EmailNotificationPreferences(user=user, frequency=EMAIL_FREQUENCY.DAILY).save()
        self.create_profiles_but_unlink_objects()
        for profile in PrisonerProfile.objects.all():
            profile.monitoring_users.add(user)
        call_command('update_security_profiles')
        call_command('send_notification_emails')

        self.assertEqual(len(mock_send_email.call_args_list), 1)
        send_email_kwargs = mock_send_email.call_args_list[0].kwargs
        self.assertEqual(send_email_kwargs['template_name'], 'api-intel-notification-first')
        self.assertTrue(user.flags.filter(name=EMAILS_STARTED_FLAG).exists())
        yesterday = timezone.now() - datetime.timedelta(days=1)
        yesterday = yesterday.date()
        transaction_count = Event.objects.filter(triggered_at__date=yesterday, user=user).count()
        self.assertEqual(send_email_kwargs['personalisation']['count'], transaction_count)
        self.assertIsNotNone(EmailNotificationPreferences.objects.get(user=user).last_sent_at)

    def test_sends_subsequent_email_with_events(self, mock_send_email):
        user = self.security_staff[0]
        user.flags.create(name=EMAILS_STARTED_FLAG)
        EmailNotificationPreferences(user=user, frequency=EMAIL_FREQUENCY.DAILY).save()
        self.create_profiles_but_unlink_objects()
        for profile in DebitCardSenderDetails.objects.all():
            profile.monitoring_users.add(user)
        call_command('update_security_profiles')
        call_command('send_notification_emails')

        self.assertEqual(len(mock_send_email.call_args_list), 1)
        send_email_kwargs = mock_send_email.call_args_list[0].kwargs
        self.assertEqual(send_email_kwargs['template_name'], 'api-intel-notification-daily')
        self.assertTrue(user.flags.filter(name=EMAILS_STARTED_FLAG).exists())
        yesterday = timezone.now() - datetime.timedelta(days=1)
        yesterday = yesterday.date()
        transaction_count = Event.objects.filter(triggered_at__date=yesterday, user=user).count()
        self.assertEqual(send_email_kwargs['personalisation']['count'], transaction_count)
        self.assertIsNotNone(EmailNotificationPreferences.objects.get(user=user).last_sent_at)

    def test_profile_grouping(self, mock_send_email):
        user = self.security_staff[0]
        call_command('update_security_profiles')
        for profile in PrisonerProfile.objects.all():
            profile.monitoring_users.add(user)
        for profile in DebitCardSenderDetails.objects.all():
            profile.monitoring_users.add(user)

        period_start = timezone.make_aware(datetime.datetime.combine(timezone.now(), datetime.time.min))
        period_start -= datetime.timedelta(days=7)
        period_end = period_start + datetime.timedelta(days=1)
        events = get_events(period_start, period_end)
        self.assertFalse(events.exists())

        prisoner_profile_1, prisoner_profile_2 = PrisonerProfile.objects.all()[:2]
        sender_profile_1, sender_profile_2 = SenderProfile.objects.filter(debit_card_details__isnull=False)[:2]
        debit_card_1 = sender_profile_1.debit_card_details.first()
        debit_card_2 = sender_profile_2.debit_card_details.first()
        credit = mommy.make(
            Credit,
            received_at=period_start, amount=100,
            prisoner_number=prisoner_profile_1.prisoner_number, prisoner_name=prisoner_profile_1.prisoner_name,
            prison=PrisonerLocation.objects.get(prisoner_number=prisoner_profile_1.prisoner_number).prison,
            resolution=CREDIT_RESOLUTION.CREDITED, reconciled=True, private_estate_batch=None,
            prisoner_profile=prisoner_profile_1, sender_profile=sender_profile_1,
            is_counted_in_prisoner_profile_total=False,
            is_counted_in_sender_profile_total=False
        )
        mommy.make(
            Payment,
            credit=credit,
            card_number_last_digits=debit_card_1.card_number_last_digits,
            card_expiry_date=debit_card_1.card_expiry_date,
            billing_address=debit_card_1.billing_addresses.first(),
        )
        credit = mommy.make(
            Credit,
            received_at=period_start, amount=200,
            prisoner_number=prisoner_profile_2.prisoner_number, prisoner_name=prisoner_profile_2.prisoner_name,
            prison=PrisonerLocation.objects.get(prisoner_number=prisoner_profile_2.prisoner_number).prison,
            resolution=CREDIT_RESOLUTION.CREDITED, reconciled=True, private_estate_batch=None,
            prisoner_profile=prisoner_profile_2, sender_profile=sender_profile_2,
            is_counted_in_prisoner_profile_total=False,
            is_counted_in_sender_profile_total=False
        )
        mommy.make(
            Payment,
            credit=credit,
            card_number_last_digits=debit_card_2.card_number_last_digits,
            card_expiry_date=debit_card_2.card_expiry_date,
            billing_address=debit_card_2.billing_addresses.first(),
        )

        call_command('update_security_profiles')

        events = get_events(period_start, period_end)
        event_group = summarise_group(group_events(events, user))
        self.assertEqual(event_group['transaction_count'], 4)
        self.assertEqual(len(event_group['senders']), 2)
        self.assertEqual(len(event_group['prisoners']), 2)

        mock_send_email.assert_not_called()

    @mock.patch('notification.management.commands.send_notification_emails.timezone')
    def test_does_not_send_email_if_already_sent_today(self, mock_timezone, mock_send_email):
        today_date = timezone.localdate()
        yesterday_date = today_date - datetime.timedelta(days=1)

        user = self.security_staff[0]
        user.flags.create(name=EMAILS_STARTED_FLAG)
        EmailNotificationPreferences(user=user, frequency=EMAIL_FREQUENCY.DAILY).save()
        self.create_profiles_but_unlink_objects()
        for profile in DebitCardSenderDetails.objects.all():
            profile.monitoring_users.add(user)
        call_command('update_security_profiles')
        self.assertTrue(Event.objects.filter(triggered_at__date=yesterday_date).exists())
        self.assertTrue(Event.objects.filter(triggered_at__date=today_date).exists())

        # pretend it's yesterday
        mock_timezone.localdate.return_value = yesterday_date
        call_command('send_notification_emails')
        self.assertEqual(len(mock_send_email.call_args_list), 1)
        self.assertEqual(EmailNotificationPreferences.objects.get(user=user).last_sent_at, yesterday_date)

        call_command('send_notification_emails')
        self.assertEqual(len(mock_send_email.call_args_list), 1)

        # now check today
        mock_timezone.localdate.return_value = today_date
        call_command('send_notification_emails')
        self.assertEqual(len(mock_send_email.call_args_list), 2)
        self.assertEqual(EmailNotificationPreferences.objects.get(user=user).last_sent_at, today_date)

        call_command('send_notification_emails')
        self.assertEqual(len(mock_send_email.call_args_list), 2)


@mock.patch('notification.management.commands.send_notification_report.send_email')
class SendNotificationReportTestCase(NotificationBaseTestCase):
    def make_2days_of_random_models(self):
        test_users = make_test_users()
        load_random_prisoner_locations()
        generate_payments(payment_batch=20, days_of_history=2)
        while not Disbursement.objects.sent().exists():
            generate_disbursements(disbursement_batch=20, days_of_history=2)
        return test_users['security_staff']

    def test_invalid_parameters(self, mock_send_email):
        with self.assertRaises(CommandError, msg='Email address should be invalid'):
            call_command('send_notification_report', 'admin')
        with self.assertRaises(CommandError, msg='Email address should be invalid'):
            call_command('send_notification_report', 'admin@mtp.local', 'admin')
        with self.assertRaises(CommandError, msg='Date should be invalid'):
            call_command('send_notification_report', 'admin@mtp.local', since='yesterday')
        with self.assertRaises(CommandError, msg='Dates should be invalid'):
            call_command('send_notification_report', 'admin@mtp.local', since='2019-08-01', until='2019-08-01')
        with self.assertRaises(CommandError, msg='Dates should be invalid'):
            call_command('send_notification_report', 'admin@mtp.local', since='2019-08-02', until='2019-08-01')
        with self.assertRaises(CommandError, msg='Date span should be invalid'):
            call_command('send_notification_report', 'admin@mtp.local', since='2019-07-01', until='2019-08-01')
        with self.assertRaises((CommandError, KeyError), msg='Rules should be invalid'):
            call_command('send_notification_report', 'admin@mtp.local', rules=['abc'])
        mock_send_email.assert_not_called()

    @mock.patch('notification.management.commands.send_notification_report.generate_report')
    def test_date_ranges(self, mock_generate_report, _mock_send_email):
        call_command('send_notification_report', 'admin@mtp.local', since='2019-08-01', until='2019-08-02')
        _workbook, period_start, period_end, _rules = mock_generate_report.call_args[0]
        self.assertEqual(period_start.date(), datetime.date(2019, 8, 1))
        self.assertEqual(period_end.date(), datetime.date(2019, 8, 2))

        today = timezone.localtime(timezone.now()).date()
        yesterday = today - datetime.timedelta(days=1)

        call_command('send_notification_report', 'admin@mtp.local', since=yesterday.isoformat())
        _workbook, period_start, period_end, _rules = mock_generate_report.call_args[0]
        self.assertEqual(period_start.date(), yesterday)
        self.assertEqual(period_end.date(), today)

        call_command('send_notification_report', 'admin@mtp.local', until=yesterday.isoformat())
        _workbook, period_start, period_end, _rules = mock_generate_report.call_args[0]
        self.assertEqual(period_start.date(), yesterday - datetime.timedelta(days=1))
        self.assertEqual(period_end.date(), yesterday)

        call_command('send_notification_report', 'admin@mtp.local')
        _workbook, period_start, period_end, _rules = mock_generate_report.call_args[0]
        self.assertEqual(period_end.date() - datetime.timedelta(days=7), period_start.date())

    def assertHasExcelAttachment(self, mock_send_email):  # noqa: N802
        self.assertEqual(len(mock_send_email.call_args_list), 1)
        send_email_kwargs = mock_send_email.call_args_list[0].kwargs
        workbook = openpyxl.load_workbook(io.BytesIO(send_email_kwargs['personalisation']['attachment']))
        return send_email_kwargs, workbook

    def test_empty_report(self, mock_send_email):
        call_command('send_notification_report', 'admin@mtp.local', since='2019-08-01', until='2019-08-02')
        send_email_kwargs, workbook = self.assertHasExcelAttachment(mock_send_email)
        self.assertIn('01 Aug 2019', send_email_kwargs['personalisation']['period_description'])
        self.assertEqual(
            len(workbook.sheetnames),
            sum(
                len(rule.applies_to_models)
                for rule in RULES.values()
            )
        )
        for worksheet in workbook.sheetnames:
            worksheet = workbook[worksheet]
            self.assertEqual(worksheet['B2'].value, 'No notifications', FLAKY_TEST_WARNING)

    def test_reports_generated(self, mock_send_email):
        self.make_2days_of_random_models()

        # move 1 credit and 1 disbursement to a past date and make them appear in HA and NWN sheets
        credit = Credit.objects.all().order_by('?').first()
        credit.received_at -= datetime.timedelta(days=7)
        credit.amount = 12501
        credit.save()
        disbursement = Disbursement.objects.sent().order_by('?').first()
        disbursement.created = credit.received_at
        disbursement.amount = 13602
        disbursement.save()

        call_command('update_security_profiles')

        # generate report
        report_date = credit.received_at.date()
        since = report_date.strftime('%Y-%m-%d')
        until = (report_date + datetime.timedelta(days=1)).strftime('%Y-%m-%d')
        call_command('send_notification_report', 'admin@mtp.local', since=since, until=until)

        send_email_kwargs, workbook = self.assertHasExcelAttachment(mock_send_email)
        self.assertIn(report_date.strftime('%d %b %Y'), send_email_kwargs['personalisation']['period_description'])

        credit_sheets = {'cred-high amount', 'cred-not whole'}
        disbursement_sheets = {'disb-high amount', 'disb-not whole'}
        expected_sheets = credit_sheets | disbursement_sheets
        self.assertTrue(expected_sheets.issubset(set(workbook.sheetnames)))

        for worksheet in credit_sheets:
            worksheet = workbook[worksheet]
            dimensions = worksheet.calculate_dimension()
            rows, _columns = coordinate_to_tuple(dimensions.split(':')[1])
            self.assertEqual(rows, 2)
            self.assertEqual(worksheet['F2'].value, '£125.01', FLAKY_TEST_WARNING)
            self.assertEqual(worksheet['H2'].value, credit.prisoner_name)
            self.assertIn(f'/credits/{credit.id}/', worksheet['B2'].hyperlink.target)
        for worksheet in disbursement_sheets:
            worksheet = workbook[worksheet]
            dimensions = worksheet.calculate_dimension()
            rows, _columns = coordinate_to_tuple(dimensions.split(':')[1])
            self.assertEqual(rows, 2)
            self.assertEqual(worksheet['F2'].value, '£136.02')
            self.assertEqual(worksheet['O2'].value, disbursement.recipient_address)
            self.assertIn(f'/disbursements/{disbursement.id}/', worksheet['B2'].hyperlink.target)

    def test_reports_generated_for_monitored_prisoners(self, mock_send_email):
        security_staff = self.make_2days_of_random_models()
        self.create_profiles_but_unlink_objects()

        # set up scenario such that every prisoner is monitored by 1 person
        # except for one prisoner that has 2 monitors
        for profile in PrisonerProfile.objects.all():
            profile.monitoring_users.add(security_staff[0])
        profile = PrisonerProfile.objects.order_by('?').first()
        profile.monitoring_users.add(security_staff[1])
        prisoner_number_with_2_monitors = profile.prisoner_number

        call_command('update_security_profiles')

        # generate reports for whole range
        report_date = Credit.objects.order_by('received_at').first().received_at.date()
        since = report_date.strftime('%Y-%m-%d')
        call_command('send_notification_report', 'admin@mtp.local', since=since)

        send_email_kwargs, workbook = self.assertHasExcelAttachment(mock_send_email)
        self.assertIn(report_date.strftime('%d %b %Y'), send_email_kwargs['personalisation']['period_description'])

        expected_sheets = {'cred-mon. prisoners', 'disb-mon. prisoners'}
        self.assertTrue(expected_sheets.issubset(set(workbook.sheetnames)))

        prisoner_number_cols = [
            ('cred-mon. prisoners', 7),
            ('disb-mon. prisoners', 7),
        ]
        for worksheet, prisoner_number_col in prisoner_number_cols:
            worksheet = workbook[worksheet]
            rows = iter(worksheet.rows)
            next(rows)
            for row in rows:
                prisoner_number = row[prisoner_number_col].value
                monitored_by = row[1].value
                if prisoner_number == prisoner_number_with_2_monitors:
                    self.assertEqual(monitored_by, 2, FLAKY_TEST_WARNING)
                else:
                    self.assertEqual(monitored_by, 1, FLAKY_TEST_WARNING)

    def test_reports_generated_for_counting_rules(self, mock_send_email):
        # make just enough credits to trigger CSFREQ rule with latest credit
        rule = RULES['CSFREQ']
        count = rule.kwargs['limit'] + 1
        sender = make_sender()
        prisoner = make_prisoner()
        credit_list = make_csfreq_credits(timezone.now(), sender, count)
        for credit in credit_list:
            credit.prisoner_profile = prisoner
            credit.save()

        # generate reports for whole range
        since = credit_list[-1].received_at.date()
        until = credit_list[0].received_at.date() + datetime.timedelta(days=1)
        call_command(
            'send_notification_report', 'admin@mtp.local',
            since=since.strftime('%Y-%m-%d'), until=until.strftime('%Y-%m-%d'),
            rules=['CSFREQ'],
        )
        _send_email_kwargs, workbook = self.assertHasExcelAttachment(mock_send_email)

        worksheet = workbook['cred-freq. source']
        dimensions = worksheet.calculate_dimension()
        rows, _columns = coordinate_to_tuple(dimensions.split(':')[1])
        self.assertEqual(rows, 2)
        self.assertEqual(worksheet['B2'].value, count)
