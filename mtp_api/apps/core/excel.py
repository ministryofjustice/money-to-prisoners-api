from django.core.files import File
import openpyxl
import xlrd


class ExcelWorkbook:
    """
    Provides a common, but very basic, read-only interface for reading XLS and XLSX files,
    based on file extension (not file contents)
    """
    subclasses = {}
    extension = NotImplemented

    def __init_subclass__(cls):
        cls.subclasses[cls.extension] = cls

    @classmethod
    def open_workbook(cls, file: File):
        extension = (file.name or '').rsplit('.')[-1].lower()
        try:
            return cls.subclasses[extension](file)
        except KeyError:
            raise TypeError('Unknown file extension')

    def __init__(self, file: File):
        self.file = file

    def __enter__(self):
        # subclasses open file here
        raise NotImplementedError

    def __exit__(self, exc_type, exc_val, exc_tb):
        # subclasses close file here
        raise NotImplementedError

    def get_sheet(self, index: int) -> 'ExcelWorksheet':
        raise NotImplementedError


class XLSWorkbook(ExcelWorkbook):
    extension = 'xls'

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._workbook: xlrd.book.Book

    def __enter__(self):
        try:
            self._workbook = xlrd.open_workbook(
                filename=self.file.name,
                file_contents=self.file.read(),
                on_demand=True,
            ).__enter__()
        except (xlrd.XLRDError, xlrd.compdoc.CompDocError) as e:
            raise TypeError('Cannot load XLS file') from e
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        return self._workbook.__exit__(exc_type, exc_val, exc_tb)

    def get_sheet(self, index: int) -> 'XLSWorksheet':
        worksheet = self._workbook.get_sheet(index)
        return XLSWorksheet(worksheet)


class XLSXWorkbook(ExcelWorkbook):
    extension = 'xlsx'

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._workbook: openpyxl.workbook.Workbook

    def __enter__(self):
        try:
            self._workbook = openpyxl.load_workbook(self.file, read_only=True)
        except ValueError as e:
            raise TypeError('Cannot load XLSX file') from e
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self._workbook.close()

    def get_sheet(self, index: int) -> 'XLSXWorksheet':
        worksheet = self._workbook.worksheets[0]
        return XLSXWorksheet(worksheet)


class ExcelWorksheet:
    """
    Returned by ExcelWorkbook.get_sheet(index),
    a common interface for dealing with XLS and XLSX worksheets
    """

    @property
    def row_count(self) -> int:
        raise NotImplementedError

    def cell_value(self, row: int, column: int):
        raise NotImplementedError


class XLSWorksheet(ExcelWorksheet):
    def __init__(self, worksheet: xlrd.sheet.Sheet):
        self._worksheet = worksheet

    @property
    def row_count(self) -> int:
        return self._worksheet.nrows

    def cell_value(self, row: int, column: int):
        return self._worksheet.cell_value(row, column)


class XLSXWorksheet(ExcelWorksheet):
    def __init__(self, worksheet: openpyxl.workbook.workbook.ReadOnlyWorksheet):
        self._worksheet = worksheet

    @property
    def row_count(self) -> int:
        return self._worksheet.max_row

    def cell_value(self, row: int, column: int):
        return self._worksheet.cell(row + 1, column + 1).value
